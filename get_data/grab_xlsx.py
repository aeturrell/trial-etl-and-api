import pandas as pd
from openpyxl import load_workbook
import xlrd
from bs4 import BeautifulSoup
from loguru import logger
import requests
from parse import parse

from pathlib import Path, PosixPath

STEM_FILENAME = "ukbusinessworkbook"
ONS_URL_MAIN = "https://www.ons.gov.uk/businessindustryandtrade/business/activitysizeandlocation/datasets/ukbusinessactivitysizeandlocation"
DOWNLOADS_LOCATION = "scratch/"
DOWNLOADED_FILENAME_STEM = "business_counts_"


def get_sheetnames_xlsx(filepath: PosixPath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames


def get_sheetnames_xls(filepath: PosixPath):
    xls = xlrd.open_workbook(filepath, on_demand=True)
    return xls.sheet_names()


def find_files(url: str):
    """Finds all xlsx and xls files at a given url.

    Args:
        url (str): Web page to look for files on.

    Returns:
        list: file links with extensions
    """
    soup = BeautifulSoup(requests.get(url).text, features="lxml")

    hrefs = [a["href"] for a in soup.find_all("a")]
    hrefs = [a for a in hrefs if len(a.split(".")) > 1]
    hrefs = [
        a for a in hrefs if (a.split(".")[1] == "xlsx" or a.split(".")[1] == "xls")
    ]
    return hrefs


def download_and_save_file(file_url: str, file_name: str):
    """Grabs known files from ONS' website and downloads them.

    Args:
        file_url (str): The url where the data file can be found.
        file_name (str): What to save the file under going forwards.

    Returns:
        str: same as input
    """
    # if scratch path doesn't exist, create it
    dl_path = Path(DOWNLOADS_LOCATION)
    dl_path.mkdir(parents=True, exist_ok=True)
    # Now check if file exists already. If not, dl it
    file_location = dl_path / file_name
    if file_location.is_file():
        logger.info(f"Skipping download of {file_name}; file already exists")
    else:
        r = requests.get("https://www.ons.gov.uk" + file_url, stream=True)
        with open(file_location, "wb") as f:
            f.write(r.content)
    logger.info(f"Success: file download of {file_name} complete")
    return file_name


list_of_hrefs = find_files(ONS_URL_MAIN)

# filter these to fit the /{year}/ukbusinessworkbook{year}.{filename} pattern
list_of_hrefs = [x for x in list_of_hrefs if STEM_FILENAME in x]
dict_of_info = [
    parse(
        "/file?uri=/businessindustryandtrade/business/activitysizeandlocation/datasets/ukbusinessactivitysizeandlocation/{year}/ukbusinessworkbook{year}.{file_extension}",
        url,
    ).named
    for url in list_of_hrefs
]

# put these in a data frame for convenience.
files_to_dl_df = pd.DataFrame.from_dict(dict_of_info)
files_to_dl_df["href"] = list_of_hrefs
files_to_dl_df["file_name"] = (
    DOWNLOADED_FILENAME_STEM
    + files_to_dl_df["year"]
    + "."
    + files_to_dl_df["file_extension"]
)

# now go through each and download it
files_to_dl_df["downloaded_files"] = files_to_dl_df.apply(
    lambda x: download_and_save_file(x["href"], x["file_name"]), axis=1
)
