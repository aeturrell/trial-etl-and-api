from openpyxl import load_workbook
import xlrd
import pandas as pd
from glob import glob

from pathlib import Path, PosixPath

DOWNLOADS_LOCATION = "scratch/"  # this should be globally defined


def get_sheetnames_xlsx(file_path: PosixPath) -> list[str]:
    """Given an xlsx file, returns the names of the sheets in it.

    Args:
        file_path (PosixPath): path to the xlsx file

    Returns:
        list[str]: A list of the worksheet names
    """
    wb = load_workbook(file_path, read_only=True, keep_links=False)
    return wb.sheetnames


def get_sheetnames_xls(filepath: PosixPath) -> list[str]:
    """Given an xls file, returns the names of the sheets in it.

    Args:
        filepath (PosixPath): path to the xls file

    Returns:
        list[str]: A list of the worksheet names
    """
    xls = xlrd.open_workbook(filepath, on_demand=True)
    return xls.sheet_names()


def get_files_list() -> pd.DataFrame():
    # parse files
    file_list = glob(DOWNLOADS_LOCATION + "[0-9][0-9][0-9][0-9].xls*")
    # create dataframe of key info, including file extension
    # extract sheetnames
    # choose likely sheetname (via closest match)
    # go through likely sheets
    # pull all complete rows, and date row
    # melt
    # stack together
    # to parquet?
    # push to cloud
